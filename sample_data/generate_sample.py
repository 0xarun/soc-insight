"""
generate_sample.py
Run this once to create a realistic sample Excel file for testing SOC Insight.
Usage: python generate_sample.py
"""

import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

random.seed(42)

ANALYSTS = ["Arun Kumar", "Priya Sharma", "Ravi Patel", "Sneha Rao", "Mohammed Ali", "Deepa Nair"]
SEVERITIES = (
    ["Critical"] * 10 + ["High"] * 20 + ["Medium"] * 40 + ["Low"] * 30
)
ALERTS = [
    "Malware Detected", "Phishing Email", "Unauthorized Access", "Brute Force Attack",
    "Data Exfiltration", "Ransomware Activity", "Suspicious Login", "Port Scan",
    "DDoS Attack", "SQL Injection", "XSS Attack", "Privilege Escalation",
    "Lateral Movement", "C2 Communication", "Insider Threat", "Vulnerability Exploit",
]
ACTIONS = [
    "Isolated host", "Blocked IP", "Reset credentials", "Patched vulnerability",
    "Alerted user", "Escalated to L3", "False positive – closed", "Quarantined file",
    "Updated firewall rules", "Notified management",
]
RESULTS = ["True Positive"] * 60 + ["False Positive"] * 40


def rand_datetime(month: int, year: int = 2026) -> datetime:
    day = random.randint(1, 28)
    hour = random.randint(7, 22)
    minute = random.randint(0, 59)
    return datetime(year, month, day, hour, minute, 0)


def fmt_td(seconds: int) -> str:
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


rows = []
inc_counter = 1000

for month in range(1, 5):   # Jan – Apr 2026
    count = random.randint(35, 60)
    for _ in range(count):
        occ = rand_datetime(month)
        sev = random.choice(SEVERITIES)
        # MTTD: critical faster detection
        mttd_secs = {
            "Critical": random.randint(5 * 60, 45 * 60),
            "High":     random.randint(10 * 60, 90 * 60),
            "Medium":   random.randint(15 * 60, 120 * 60),
            "Low":      random.randint(20 * 60, 180 * 60),
        }[sev]
        mttr_secs = {
            "Critical": random.randint(1 * 3600, 8 * 3600),
            "High":     random.randint(2 * 3600, 12 * 3600),
            "Medium":   random.randint(1 * 3600, 6 * 3600),
            "Low":      random.randint(30 * 60, 3 * 3600),
        }[sev]
        det = occ + timedelta(seconds=mttd_secs)
        rows.append({
            "INC Number":       f"INC-{inc_counter}",
            "Date":             occ.strftime("%d/%m/%Y"),
            "Analyst":          random.choice(ANALYSTS),
            "Alert":            random.choice(ALERTS),
            "Severity":         sev,
            "Action Taken":     random.choice(ACTIONS),
            "Occurrence Time":  occ.strftime("%d/%m/%Y %H:%M:%S"),
            "Detection Time":   det.strftime("%d/%m/%Y %H:%M:%S"),
            "MTTD":             fmt_td(mttd_secs),
            "MTTR":             fmt_td(mttr_secs),
            "Result":           random.choice(RESULTS),
        })
        inc_counter += 1

# Write Excel with nice formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Incidents"

headers = list(rows[0].keys())
header_fill = PatternFill("solid", fgColor="1e3a5f")
header_font = Font(color="FFFFFF", bold=True, size=11)

ws.append(headers)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row in rows:
    ws.append(list(row.values()))

# Column widths
col_widths = [14, 12, 16, 22, 10, 22, 22, 22, 10, 10, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = w

out_path = Path(__file__).parent / "sample_incidents.xlsx"
wb.save(out_path)
print(f"✅ Sample data saved: {out_path}")
print(f"   Rows generated: {len(rows)}")
